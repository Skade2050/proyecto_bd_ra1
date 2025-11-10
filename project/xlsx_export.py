from __future__ import annotations
from pathlib import Path
import pandas as pd

# Columnas mínimas esperadas (permite extras)
EXPECTED_COLS = [
    "fecha", "id_respuesta", "canal", "producto",
    "satisfaccion", "comentario", "tienda", "agente",
]

def _leer_csv(path: str) -> pd.DataFrame:
    """Lee CSV con detección simple de separador. Rechaza vacíos o con <2 columnas."""
    for sep in (";", ",", "\t", "|"):
        try:
            df = pd.read_csv(path, sep=sep)
            # exigimos al menos 2 columnas y que no esté vacío
            if not df.empty and df.shape[1] >= 2:
                return df
        except Exception:
            continue
    raise ValueError("No se pudo detectar separador o el CSV está vacío. Guarda el CSV con ';' o ','.")

def _normalizar(df: pd.DataFrame) -> pd.DataFrame:
    """Ajusta columnas, tipos y textos para el Excel del trabajo."""
    faltan = [c for c in EXPECTED_COLS if c not in df.columns]
    if faltan:
        raise ValueError(f"Faltan columnas obligatorias: {faltan}")

    # Reordenar: primero las esperadas, luego las extra
    df = df[[c for c in EXPECTED_COLS if c in df.columns] + [c for c in df.columns if c not in EXPECTED_COLS]]

    # Trim textos (sin convertir NaN en "nan")
    for c in df.columns:
        if pd.api.types.is_object_dtype(df[c]) or pd.api.types.is_string_dtype(df[c]):
            df[c] = df[c].astype("string")
            df[c] = df[c].str.strip()

    # fecha (day-first), id como texto (NA preservado), satisfaccion numérica (coma o punto)
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce", dayfirst=True)

    # preservar nulos y castear a dtype string (no a str nativo para evitar "nan")
    df["id_respuesta"] = df["id_respuesta"].astype("string")

    df["satisfaccion"] = pd.to_numeric(
        df["satisfaccion"].astype("string").str.replace(",", ".", regex=False).str.replace(" ", "", regex=False),
        errors="coerce"
    )
    return df

def _autosize(ws, df: pd.DataFrame):
    """Auto-ajusta ancho de columnas (openpyxl)."""
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(df.columns, start=1):
        texto = df[col].astype(str).head(1000)
        ancho = max(len(str(col)), *(len(x) for x in texto)) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(ancho, 60)

def _escribir_hojas(xlsx_path: str, df: pd.DataFrame):
    """Crea 'Encuestas' y 'ResumenCalidad' con formato básico."""
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        # Hoja principal
        df.to_excel(w, sheet_name="Encuestas", index=False)
        ws = w.book["Encuestas"]
        from openpyxl.styles import Font
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize(ws, df)

        # ResumenCalidad
        q = pd.DataFrame({
            "columna": df.columns,
            "dtype": df.dtypes.astype(str).values,
            "n_nulos": df.isna().sum().values,
            "%_nulos": (df.isna().sum() / max(len(df), 1) * 100).round(2).values,
            "n_distintos": [df[c].nunique(dropna=True) for c in df.columns],
            "muestra_valores": ["; ".join(df[c].astype(str).head(3).tolist()) for c in df.columns],
        })
        q.to_excel(w, sheet_name="ResumenCalidad", index=False)
        wsq = w.book["ResumenCalidad"]
        for c in wsq[1]:
            c.font = Font(bold=True)
        wsq.freeze_panes = "A2"
        wsq.auto_filter.ref = wsq.dimensions
        _autosize(wsq, q)

def build_xlsx_from_csv(csv_path: str, xlsx_path: str | None = None) -> str:
    """CSV → XLSX (Encuestas + ResumenCalidad). Devuelve ruta del XLSX."""
    src = Path(csv_path)
    if not src.is_file():
        raise FileNotFoundError(f"No existe el CSV: {csv_path}")

    df = _leer_csv(str(src))
    df = _normalizar(df)

    out = Path(xlsx_path) if xlsx_path else Path("project/output/xlsx") / (src.stem + ".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    _escribir_hojas(str(out), df)
    return str(out)

def build_xlsx_from_df(df: pd.DataFrame, xlsx_path: str) -> str:
    """DataFrame → XLSX (mismo formato). Devuelve ruta del XLSX."""
    df = _normalizar(df.copy())
    out = Path(xlsx_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    _escribir_hojas(str(out), df)
    return str(out)
